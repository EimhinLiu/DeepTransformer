from sklearn.metrics import roc_auc_score, precision_recall_curve, auc, roc_curve
import numpy as np
import torch
import random
import openpyxl
import pandas as pd
from rdkit import Chem
from rdkit.Chem import AllChem
from rdkit import DataStructs
from pubchempy import Compound
import os
import matplotlib.pyplot as plt

def rocauc(input_dict):
    y_true, y_pred = input_dict['y_true'], input_dict['y_pred']
    
    y_true = y_true.detach().cpu().numpy()
    y_pred = y_pred.detach().cpu().numpy()
    
    rocauc_list = []
    for i in range(y_true.shape[1]):
        if np.sum(y_true[:,i] == 1) > 0 and np.sum(y_true[:,i] == 0) > 0:
            is_labeled = y_true[:,i] == y_true[:,i]
            rocauc_list.append(roc_auc_score(y_true[is_labeled,i], y_pred[is_labeled,i]))

    return {'rocauc': sum(rocauc_list)/len(rocauc_list)}

def roccurve(input_dict):
    y_true, y_pred = input_dict['y_true'], input_dict['y_pred']
    
    y_true = y_true.detach().cpu().numpy()
    y_pred = y_pred.detach().cpu().numpy()
    
    for i in range(y_true.shape[1]):
        if np.sum(y_true[:,i] == 1) > 0 and np.sum(y_true[:,i] == 0) > 0:
            is_labeled = y_true[:,i] == y_true[:,i]
            fpr, tpr, thresholds = roc_curve(y_true[is_labeled,i], y_pred[is_labeled,i])
    
    return fpr, tpr

def prauc(input_dict):
    y_true, y_pred = input_dict['y_true'], input_dict['y_pred']
    
    y_true = y_true.detach().cpu().numpy()
    y_pred = y_pred.detach().cpu().numpy()
    
    prauc_list = []
    for i in range(y_true.shape[1]):
        if np.sum(y_true[:,i] == 1) > 0 and np.sum(y_true[:,i] == 0) > 0:
            is_labeled = y_true[:,i] == y_true[:,i]
            precision, recall, thresholds = precision_recall_curve(y_true[is_labeled,i], y_pred[is_labeled,i])
            prauc_list.append(auc(recall, precision))

    return {'prauc': sum(prauc_list)/len(prauc_list)}

def prcurve(input_dict):
    y_true, y_pred = input_dict['y_true'], input_dict['y_pred']
    
    y_true = y_true.detach().cpu().numpy()
    y_pred = y_pred.detach().cpu().numpy()
    
    for i in range(y_true.shape[1]):
        if np.sum(y_true[:,i] == 1) > 0 and np.sum(y_true[:,i] == 0) > 0:
            is_labeled = y_true[:,i] == y_true[:,i]
            precision, recall, thresholds = precision_recall_curve(y_true[is_labeled,i], y_pred[is_labeled,i])
    
    return precision, recall

def multi_curve_drawing(paths):
    au = pd.read_csv(r'.\all_model_results.csv')
    au = np.array(au)
    
    metric = paths[0].split('_')[0].split('\\')[2]
    fig, ax = plt.subplots()
    
    for path in paths:
        auc = 0
        aupr = 0
        if 'res+' in path:
            if metric == 'roc':
                ax.plot([0,1], [0,1], linestyle = '--', color = 'grey')
                plt.xlabel('False Positive Rate')
                plt.ylabel('True Positive Rate')
                model = path.split('roc_')[1].split('.')[0]
                for a in au:
                    m = a[0]
                    if model == m:
                        auc = round(a[1], 4)
            elif metric == 'pr':
                ax.plot([0,1], [1,0], linestyle = '--', color = 'grey')
                plt.xlabel('Recall')
                plt.ylabel('Precision')
                model = path.split('pr_')[1].split('.')[0]
                for a in au:
                    m = a[0]
                    if model == m:
                        aupr = round(a[2], 4)
                    
            model_name = path.split('_')[2].split('Conv')[0]
            if model_name == 'General':
                model_name = 'GraphGym'
            elif model_name == 'NN':
                model_name = 'MPNN'
            elif model_name == 'Transformer':
                model_name = 'GraphTransformer'
            
            x = sum(np.array(pd.read_csv(path, usecols=[0])).tolist(), [])
            y = sum(np.array(pd.read_csv(path, usecols=[1])).tolist(), [])
            if 'TransformerConv' in path:
                if metric == 'roc':
                    ax.plot(x, y, label = model_name + f'(auc={str(auc)})', color = 'gold')
                elif metric == 'pr':
                    ax.plot(x, y, label = model_name + f'(aupr={str(aupr)})', color = 'gold')
            else:
                if metric == 'roc':
                    ax.plot(x, y, label = model_name + f'(auc={str(auc)})')
                elif metric == 'pr':
                    ax.plot(x, y, label = model_name + f'(aupr={str(aupr)})')
            plt.legend(fontsize=7)
        
    fig.savefig(rf'./drawing/{metric}_multi_curve.png', dpi = 300)

def acc(input_dict):
    y_true, y_pred = input_dict['y_true'], input_dict['y_pred']
    
    y_true = y_true.detach().cpu().numpy()
    y_pred = y_pred.detach().cpu().numpy()
    
    acc_list = []
    for i in range(y_true.shape[1]):
        is_labeled = y_true[:,i] == y_true[:,i]
        correct = y_true[is_labeled,i] == y_pred[is_labeled,i]
        acc_list.append(float(np.sum(correct))/len(correct))

    return {'acc': sum(acc_list)/len(acc_list)}

class EarlyStopping:
    """Early stops the training if validation loss doesn't improve after a given patience."""
    def __init__(self, patience=7, verbose=False, delta=0, path='checkpoint.pt', trace_func=print):
        """
        Args:
            patience (int): How long to wait after last time validation loss improved.
                            Default: 7
            verbose (bool): If True, prints a message for each validation loss improvement. 
                            Default: False
            delta (float): Minimum change in the monitored quantity to qualify as an improvement.
                            Default: 0
            path (str): Path for the checkpoint to be saved to.
                            Default: 'checkpoint.pt'
            trace_func (function): trace print function.
                            Default: print            
        """
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.val_loss_min = np.Inf
        self.delta = delta
        self.path = path
        self.trace_func = trace_func

    def __call__(self, val_loss, model):
        score = -val_loss

        if self.best_score is None:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
        elif score < self.best_score + self.delta:
            self.counter += 1
            # self.trace_func(f'EarlyStopping counter: {self.counter} out of {self.patience}')
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
            self.counter = 0

    def save_checkpoint(self, val_loss, model):
        '''Saves model when validation loss decrease.'''
        # if self.verbose:
        #     self.trace_func(f'Validation loss decreased ({self.val_loss_min:.6f} --> {val_loss:.6f}).  Saving model ...')
        torch.save(model.state_dict(), self.path)
        self.val_loss_min = val_loss

def setup_seed(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    np.random.seed(seed)
    random.seed(seed)
    torch.backends.cudnn.deterministic = True
    
def save_result(metrics, path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    first_line = ['Epoch', 'Train_Loss', 'Valid_Loss', 'Test_Rocauc', 'Test_Prauc', 'Test_Variance']
    for fir in range(len(first_line)):
        worksheet.cell(1, fir + 1, first_line[fir])
    
    for mets in range(len(metrics)):
        for met in range(len(metrics[mets])):
            worksheet.cell(mets + 2, met + 1, metrics[mets][met])
    
    workbook.save(filename=path)

def read_result(path):
    data = pd.read_excel(path, usecols = ['Test_Rocauc', 'Test_Prauc'])
    values = data.values.tolist()
    
    total_Rocauc = 0
    total_Prauc = 0
    for val in values[-10:]:
        total_Rocauc += val[0]
        total_Prauc += val[1]
    
    average_Rocauc = round(total_Rocauc / 10, 7)
    average_Prauc = round(total_Prauc / 10, 7)
    
    return average_Rocauc, average_Prauc

def calculate_chemical_similarity(input_path, output_path):
    """
    Need getting the compound CID from Pubchem database.
    """
    cid = []
    smiles = []
    with open(input_path, 'r') as f:
        data = f.read().split(' ')
        for i in range(len(data)):
            cid.append(data[i].strip())
            smile = Compound.from_cid(int(data[i].strip())).canonical_smiles
            smiles.append(smile)
    
    fps = []
    for smi in smiles:
        mol = Chem.MolFromSmiles(smi)
        fp = AllChem.GetMACCSKeysFingerprint(mol)
        fps.append(fp)
    
    chemical1, chemical2 = [], []
    for i in range(len(fps)-1):
        similarity = DataStructs.BulkTanimotoSimilarity(fps[i], fps[i+1:]) # Compare one by one with the next one to the last one.
        for s in range(len(similarity)):
            if similarity[s] >= 0.7:
                chemical1.append(cid[i])
                chemical2.append(cid[i+1+s])
                
    chemical_similarity_data = pd.DataFrame({'chemical1':chemical1, 'chemical2':chemical2})
    chemical_similarity_data.to_csv(output_path, header=False, index = False)
    
if __name__ == '__main__':
    # input_path = r'.\MolecularDockingExperience\SDF_CID.txt'
    # output_path = r'.\data\Chemical_Similar.csv'
    # calculate_chemical_similarity(input_path, output_path)

    # skip_connection = ['res+', 'res', 'plain']
    # GNNConv = ['GENConv', 'GATConv', 'GATv2Conv', 'TransformerConv', 'GeneralConv', 'GMMConv', 'NNConv']
    # conv_layers = ['30', '29', '28', '27', '26', '25', '24', '23', '22', '21', '20', '3']
    # models = []
    # average_Rocaucs = []
    # average_Praucs = []
    # for sc in skip_connection:
    #     for g in GNNConv:
    #         for cl in conv_layers:
    #             name = cl + '_' + g + '_' + sc
    #             path = os.path.join('.\metrics', name + '.xlsx')
    #             if os.path.exists(path):
    #                 average_Rocauc, average_Prauc = read_result(path)
    #                 models.append(name)
    #                 average_Rocaucs.append(average_Rocauc)
    #                 average_Praucs.append(average_Prauc)
    # all_model_results = pd.DataFrame({'Model':models, 'AverageRocauc':average_Rocaucs, 'AveragePrauc':average_Praucs})
    # all_model_results.to_csv(r".\all_model_results.csv", index = False)
    
    # pr_paths = []
    # roc_paths = []
    # for filewalks in os.walk(r'.\drawing'):
    #     for files in filewalks[2]:
    #         if 'pr' in files:
    #             pr_paths.append(os.path.join('.\\drawing', files))
    #         if 'roc' in files:
    #             roc_paths.append(os.path.join('.\\drawing', files))
    # multi_curve_drawing(pr_paths)
    # multi_curve_drawing(roc_paths)
    
    pass